from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
)

from app.models.registration import Registration


class RegistrationExportService:

    HEADERS = [
        "ID",
        "Event",
        "First Name",
        "Last Name",
        "Gender",
        "Phone",
        "Email",
        "Country",
        "State",
        "City",
        "Denomination",
        "Other Denomination",
        "Accommodation",
        "Registered At",
    ]

    @staticmethod
    def _rows(registrations):
        rows = []

        for registration in registrations:
            event_title = (
                registration.event.title
                if registration.event
                else "Unknown Event"
            )

            rows.append([
                registration.id,
                event_title,
                registration.first_name,
                registration.last_name,
                registration.gender,
                registration.phone,
                registration.email or "",
                registration.country,
                registration.state,
                registration.city,
                registration.denomination,
                registration.other_denomination or "",
                "Yes" if registration.accommodation else "No",
                registration.created_at.strftime(
                    "%Y-%m-%d %H:%M"
                ),
            ])

        return rows

    @staticmethod
    def to_excel(registrations) -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Registrations"

        # Header
        for column, header in enumerate(
            RegistrationExportService.HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=1,
                column=column,
                value=header,
            )

            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="1F3875",
            )
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )
            cell.alignment = Alignment(
                horizontal="center"
            )

        # Data
        rows = RegistrationExportService._rows(
            registrations
        )

        for row in rows:
            worksheet.append(row)

        # Column widths
        widths = [
            8,   # ID
            25,  # Event
            18,  # First name
            18,  # Last name
            12,  # Gender
            18,  # Phone
            30,  # Email
            15,  # Country
            18,  # State
            18,  # City
            22,  # Denomination
            22,  # Other denomination
            18,  # Accommodation
            22,  # Registered at
        ]

        for index, width in enumerate(
            widths,
            start=1,
        ):
            worksheet.column_dimensions[
                chr(64 + index)
            ].width = width

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    @staticmethod
    def to_pdf(registrations) -> BytesIO:
        output = BytesIO()

        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()

        title = Paragraph(
            "Registration Report",
            styles["Title"],
        )

        headers = RegistrationExportService.HEADERS

        data = [
            headers,
            *RegistrationExportService._rows(
                registrations
            ),
        ]

        table = Table(
            data,
            repeatRows=1,
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1F3875"),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    6,
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#F4F6FA"),
                    ],
                ),
            ])
        )

        document.build([
            title,
            table,
        ])

        output.seek(0)

        return output